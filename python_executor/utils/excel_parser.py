"""
Excel 解析工具 - 用于从 Excel 文件中解析用例映射数据
使用 openpyxl（纯 Python）替代 pandas 以简化 PyInstaller 打包
"""
import os
import tempfile
import uuid
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from utils.logger import get_logger
from core.case_mapping_manager import get_case_mapping_manager

logger = get_logger("excel_parser")


class CaseMappingExcelParser:
    """用例映射 Excel 解析器"""

    SYSTEM_FIELDS = [
        ('case_no', 'Case编号', True),
        ('case_name', '用例名称', True),
        ('category', '分类', True),
        ('module', '模块名称', False),
        ('script_path', '脚本路径', False),
        ('priority', '优先级', False),
        ('version', '版本号', False),
        ('tags', '标签', False),
        ('description', '描述', False),
        ('enabled', '是否启用', False)
    ]

    SYSTEM_FIELD_NAMES = [f[0] for f in SYSTEM_FIELDS]

    VALID_CATEGORIES = ['system', 'canoe', 'tsmaster', 'ttman', 'config', 'task']

    FIELD_DISPLAY_NAMES = {f[0]: f[1] for f in SYSTEM_FIELDS}

    FIELD_HINTS = {
        'case_no': '唯一标识，如 CANOE-001',
        'case_name': '用例显示名称',
        'category': f"可选值: {', '.join(VALID_CATEGORIES)}",
        'module': '所属模块名称',
        'script_path': '脚本文件路径',
        'priority': '数字越大优先级越高',
        'version': '版本号，默认1.0',
        'tags': '逗号分隔的标签',
        'description': '用例描述',
        'enabled': 'TRUE启用，FALSE禁用'
    }

    @classmethod
    def is_available(cls) -> bool:
        """检查依赖是否满足"""
        return OPENPYXL_AVAILABLE

    @classmethod
    def _get_cell_value(cls, cell) -> str:
        """获取单元格值并转换为字符串"""
        if cell.value is None:
            return ''
        return str(cell.value)

    @classmethod
    def parse_file(cls, file_path: str, sheet_index: int = 0) -> Dict[str, Any]:
        """解析 Excel 文件

        Args:
            file_path: Excel 文件路径
            sheet_index: 工作表索引（从0开始）

        Returns:
            {
                'headers': List[str],  # 表头
                'data': List[List],     # 数据行
                'row_count': int,       # 总行数
                'sheet_names': List[str],  # 所有工作表名
                'current_sheet': int      # 当前工作表索引
            }
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed, run: pip install openpyxl")

        try:
            wb = load_workbook(file_path, data_only=True)
            sheet_names = wb.sheetnames

            if sheet_index >= len(sheet_names):
                sheet_index = 0

            ws = wb[sheet_names[sheet_index]]

            headers = []
            data = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    headers = [str(cell) if cell is not None else '' for cell in row]
                else:
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    if any(cell for cell in row_data):
                        data.append(row_data)

            return {
                'headers': headers,
                'data': data,
                'row_count': len(data),
                'sheet_names': sheet_names,
                'current_sheet': sheet_index
            }
        except Exception as e:
            logger.error(f"解析 Excel 文件失败: {e}")
            raise ValueError(f"无法解析 Excel 文件: {e}")

    @classmethod
    def get_sheets_info(cls, file_path: str) -> List[Dict[str, Any]]:
        """获取 Excel 所有工作表信息

        Args:
            file_path: Excel 文件路径

        Returns:
            [{'index': 0, 'name': 'Sheet1', 'row_count': 100}, ...]
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not installed")

        try:
            wb = load_workbook(file_path, data_only=True)
            sheets_info = []

            for idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    if any(cell for cell in row):
                        row_count += 1
                sheets_info.append({
                    'index': idx,
                    'name': sheet_name,
                    'row_count': row_count
                })

            return sheets_info
        except Exception as e:
            logger.error(f"获取工作表信息失败: {e}")
            raise ValueError(f"无法获取工作表信息: {e}")

    @classmethod
    def get_headers(cls, file_path: str) -> List[str]:
        """获取 Excel 表头"""
        result = cls.parse_file(file_path)
        return result['headers']

    @classmethod
    def get_preview_data(cls, file_path: str, max_rows: int = 5, sheet_index: int = 0) -> List[Dict[str, str]]:
        """获取预览数据

        Args:
            file_path: Excel 文件路径
            max_rows: 最大预览行数
            sheet_index: 工作表索引（从0开始）

        Returns:
            预览数据列表
        """
        result = cls.parse_file(file_path, sheet_index=sheet_index)
        headers = result['headers']
        data = result['data'][:max_rows]

        preview = []
        for row in data:
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell)
            preview.append(row_dict)

        return preview

    @classmethod
    def apply_mapping(cls, file_path: str, column_mapping: Dict[str, str],
                     default_category: str = None, batch_script_path: str = None,
                     sheet_index: int = 0) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """根据列映射转换数据

        Args:
            file_path: Excel 文件路径
            column_mapping: {Excel列名: 系统字段名}
            default_category: 默认分类（当未映射category时使用）
            batch_script_path: 批量设置的脚本路径前缀
            sheet_index: 工作表索引

        Returns:
            (有效数据列表, 无效数据及错误列表)
        """
        result = cls.parse_file(file_path, sheet_index=sheet_index)
        headers = result['headers']
        data = result['data']

        valid_mappings = []
        invalid_mappings = []

        hint_keywords = ['可选项', '必填项', '必填', '选项', '说明', '提示']

        for row_idx, row in enumerate(data, start=2):
            row_dict = {}
            row_dict['_row'] = row_idx
            row_dict['_batch_script_path'] = batch_script_path

            first_cell_value = str(row[0]).strip() if len(row) > 0 else ''
            if any(keyword in first_cell_value for keyword in hint_keywords):
                continue

            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    header = headers[col_idx]
                    if header in column_mapping:
                        system_field = column_mapping[header]
                        if system_field not in row_dict or not row_dict[system_field]:
                            row_dict[system_field] = str(cell).strip()

            if default_category and (not row_dict.get('category') or row_dict.get('category') == ''):
                row_dict['category'] = default_category

            is_valid, errors = cls._validate_mapping(row_dict)
            if is_valid:
                valid_mappings.append(row_dict)
            else:
                invalid_mappings.append({
                    'row': row_idx,
                    'data': row_dict,
                    'errors': errors
                })

        return valid_mappings, invalid_mappings

    @classmethod
    def _validate_mapping(cls, row_dict: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """验证单条映射数据"""
        errors = []

        case_no = row_dict.get('case_no', '').strip()
        if not case_no:
            errors.append('case_no 不能为空')

        case_name = row_dict.get('case_name', '').strip()
        if not case_name:
            errors.append('case_name 不能为空')

        category = row_dict.get('category', '').strip()
        if not category:
            errors.append('category 不能为空')
        elif category.lower() not in [c.lower() for c in cls.VALID_CATEGORIES]:
            errors.append(f'category 值非法，可选值: {", ".join(cls.VALID_CATEGORIES)}')

        priority_str = row_dict.get('priority', '').strip()
        if priority_str:
            try:
                int(priority_str)
            except ValueError:
                errors.append('priority 必须是整数')

        enabled_str = row_dict.get('enabled', '').strip()
        if enabled_str and enabled_str.upper() not in ['TRUE', 'FALSE', '1', '0', '']:
            errors.append('enabled 必须是 TRUE/FALSE/1/0')

        return len(errors) == 0, errors

    @classmethod
    def convert_mappings_for_import(cls, valid_mappings: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """转换映射数据为导入格式

        Args:
            valid_mappings: 有效映射列表

        Returns:
            用于导入的标准化数据列表
        """
        converted = []

        for mapping in valid_mappings:
            item = {}

            item['case_no'] = mapping.get('case_no', '').strip()

            item['case_name'] = mapping.get('case_name', '').strip()

            item['category'] = mapping.get('category', '').strip().lower()

            item['module'] = mapping.get('module', '').strip()

            batch_script_path = mapping.get('_batch_script_path')
            script_path = mapping.get('script_path', '').strip()
            if batch_script_path:
                if script_path:
                    if script_path.startswith('/') or script_path.startswith('\\') or script_path[1:2] == ':':
                        item['script_path'] = script_path
                    else:
                        item['script_path'] = batch_script_path.rstrip('/\\') + '/' + script_path
                else:
                    item['script_path'] = batch_script_path
            else:
                item['script_path'] = script_path

            priority_str = mapping.get('priority', '').strip()
            item['priority'] = int(priority_str) if priority_str else 0

            item['version'] = mapping.get('version', '').strip() or '1.0'

            tags_str = mapping.get('tags', '').strip()
            item['tags'] = [t.strip() for t in tags_str.split(',') if t.strip()] if tags_str else []

            item['description'] = mapping.get('description', '').strip()

            enabled_str = mapping.get('enabled', '').strip()
            if enabled_str.upper() in ['TRUE', '1']:
                item['enabled'] = True
            elif enabled_str.upper() in ['FALSE', '0']:
                item['enabled'] = False
            else:
                item['enabled'] = True

            # 计算 ini_config_preview 并同时保存到 ini_config
            case_no = item.get('case_no', '')
            category = item.get('category', '')
            if case_no and category:
                try:
                    mapping_manager = get_case_mapping_manager()
                    ini_config_preview = mapping_manager.apply_ini_config_rule(case_no, category)
                    item['ini_config_preview'] = ini_config_preview
                    item['ini_config'] = ini_config_preview  # 保存自动计算的ini_config
                except Exception as e:
                    logger.warning(f"计算ini_config_preview失败: {e}")
                    item['ini_config_preview'] = ''
                    item['ini_config'] = ''
            else:
                item['ini_config_preview'] = ''
                item['ini_config'] = ''

            converted.append(item)

        return converted

    @classmethod
    def create_template(cls, output_path: str = None) -> str:
        """创建导入模板 Excel 文件

        Args:
            output_path: 输出路径，默认在 data 目录下

        Returns:
            模板文件路径
        """
        if output_path is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            output_path = os.path.join(base_dir, 'data', 'case_mapping_template.xlsx')

        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        if OPENPYXL_AVAILABLE:
            return cls._create_template_with_openpyxl(output_path)
        else:
            return cls._create_template_as_csv(output_path)

    @classmethod
    def _create_template_as_csv(cls, output_path: str) -> str:
        """创建 CSV 格式的模板（无 Excel 库时的备选方案）"""
        import csv

        headers = [f[1] for f in cls.SYSTEM_FIELDS]

        example_data = [
            ['TG1_TC01', 'CANoe发送报文测试', 'canoe', 'CANoe功能测试', '', '1', '1.0', '报文测试,CANoe', '测试CANoe发送报文功能', 'TRUE'],
            ['TG1_TC02', 'CANoe接收报文测试', 'canoe', 'CANoe功能测试', '', '1', '1.0', '报文测试,CANoe', '测试CANoe接收报文功能', 'TRUE'],
            ['TG1_TC03', 'CANoe信号测量', 'canoe', 'CANoe功能测试', '', '1', '1.0', '信号测试,CANoe', '测试CANoe信号测量功能', 'TRUE'],
            ['TS_SIG_01', 'TSMaster信号读取', 'tsmaster', 'TSMaster功能测试', '', '1', '1.0', '信号测试,TSMaster', '测试TSMaster信号读取', 'TRUE'],
            ['TS_SIG_02', 'TSMaster信号写入', 'tsmaster', 'TSMaster功能测试', '', '1', '1.0', '信号测试,TSMaster', '测试TSMaster信号写入', 'TRUE'],
        ]

        csv_path = output_path.replace('.xlsx', '.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(example_data)

        logger.warning(f"openpyxl 未安装，创建 CSV 模板代替: {csv_path}")
        logger.info(f"如需 Excel 模板，请安装: pip install openpyxl")
        return csv_path

    @classmethod
    def _create_template_with_openpyxl(cls, output_path: str) -> str:
        """使用 openpyxl 创建模板"""
        wb = Workbook()
        ws = wb.active
        ws.title = '用例映射导入模板'

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        required_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

        headers = [f[1] for f in cls.SYSTEM_FIELDS]
        system_fields = [f[0] for f in cls.SYSTEM_FIELDS]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

            field_name = system_fields[col_idx - 1]
            is_required = next((f[2] for f in cls.SYSTEM_FIELDS if f[0] == field_name), False)
            if is_required:
                ws.cell(row=2, column=col_idx).fill = required_fill

        example_data = [
            ['CANOE-001', 'CANoe安装路径检查', 'canoe', 'CANoe测试', 'core/functional_test_runner.py', '1', '1.0', '环境检查,CANoe', '检查CANoe安装路径是否存在', 'TRUE'],
            ['CANOE-002', 'CANoe可执行文件检查', 'canoe', 'CANoe测试', 'core/functional_test_runner.py', '1', '1.0', '环境检查,CANoe', '检查CANoe64.exe是否存在', 'TRUE'],
            ['SYS-001', 'Python环境检查', 'system', '系统环境测试', 'core/functional_test_runner.py', '1', '1.0', '环境检查,Python', '检查Python版本和必要模块', 'TRUE'],
            ['TS-001', 'TSMaster安装路径检查', 'tsmaster', 'TSMaster测试', 'core/functional_test_runner.py', '1', '1.0', '环境检查,TSMaster', '检查TSMaster安装路径', 'TRUE'],
        ]

        for row_idx, row_data in enumerate(example_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        category_col_idx = 3
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(
            type='list',
            formula1='"' + ','.join(cls.VALID_CATEGORIES) + '"',
            allow_blank=True
        )
        dv.error = '请选择有效值'
        dv.errorTitle = '无效值'
        dv.prompt = '请从下拉列表选择'
        dv.promptTitle = '分类'
        ws.add_data_validation(dv)
        dv.add(f'{get_column_letter(category_col_idx)}3:{get_column_letter(category_col_idx)}1000')

        enabled_col_idx = 10
        dv_enabled = DataValidation(
            type='list',
            formula1='"TRUE,FALSE"',
            allow_blank=True
        )
        ws.add_data_validation(dv_enabled)
        dv_enabled.add(f'{get_column_letter(enabled_col_idx)}3:{get_column_letter(enabled_col_idx)}1000')

        col_widths = [15, 25, 12, 15, 30, 10, 10, 20, 30, 10]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 25

        hint_fill = PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
        hint_font = Font(size=9, italic=True, color='666666')

        for col_idx, field_name in enumerate(system_fields, start=1):
            hint_cell = ws.cell(row=len(example_data) + 3, column=col_idx)
            hint_cell.value = cls.FIELD_HINTS.get(field_name, '')
            hint_cell.fill = hint_fill
            hint_cell.font = hint_font

        wb.save(output_path)
        logger.info(f"创建导入模板: {output_path}")

        return output_path

    @classmethod
    def save_upload_file(cls, file_stream, original_filename: str) -> str:
        """保存上传的文件到临时目录

        Args:
            file_stream: 文件流
            original_filename: 原始文件名

        Returns:
            保存后的文件路径
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        random_suffix = uuid.uuid4().hex[:8]
        ext = os.path.splitext(original_filename)[1] or '.xlsx'

        filename = f'case_mapping_import_{timestamp}_{random_suffix}{ext}'

        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, filename)

        with open(output_path, 'wb') as f:
            f.write(file_stream.read() if hasattr(file_stream, 'read') else file_stream)

        logger.info(f"保存上传文件: {output_path}")
        return output_path

    @classmethod
    def cleanup_temp_file(cls, file_path: str):
        """清理临时文件"""
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.debug(f"清理临时文件: {file_path}")
        except Exception as e:
            logger.warning(f"清理临时文件失败: {e}")
